#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
巴菲特安全边际法筛选 - 渲染脚本
读取 _raw_<TRADE_DATE>.json（由技能执行阶段逐步填充），
输出 HTML 表格 与 Excel(.xlsx) 双格式文件，内容完全一致。

用法:
  python3 buffett_screen.py --input _raw_2026-07-21.json \
       --output-dir ./outputs --date 2026-07-21

JSON 结构见文件底部 EXAMPLE 注释。
"""
import argparse
import json
import os
import datetime

# ---------- 列定义（顺序固定，HTML 与 XLSX 共用）----------
# (表头, json字段, 类型)
# 类型: no | text | pct | num2 | score | advice
COLUMNS = [
    ("序号", "no", "no"),
    ("股票代码", "code", "text"),
    ("股票名称", "name", "text"),
    ("所属行业/板块", "industry", "text"),
    ("ROE(5年均值)", "roe_5y", "pct"),
    ("毛利率(5年均值)", "gross_margin_5y", "pct"),
    ("PE(TTM)", "pe_ttm", "num2"),
    ("PEG", "peg", "num2"),
    ("资产负债率", "debt_ratio", "pct"),
    ("经营现金流/净利润", "ocf_netprofit", "num2"),
    ("盈利能力得分", "score_profit", "score"),
    ("护城河得分", "score_moat", "score"),
    ("估值合理得分", "score_valuation", "score"),
    ("财务稳健得分", "score_solvency", "score"),
    ("盈利质量得分", "score_quality", "score"),
    ("补充因子加分", "bonus_factor", "num2"),
    ("加权总分", "weighted_total", "num2"),
    ("内在价值估算(元)", "intrinsic_value", "num2"),
    ("当前股价(元)", "price", "num2"),
    ("安全边际率", "margin_of_safety", "pct"),
    ("操作建议", "advice", "advice"),
    ("风险提示", "risk_note", "text"),
    ("数据完整度", "data_completeness", "text"),
]

ADVICE_GREEN = {"重仓买入", "分批买入"}
ADVICE_ORANGE = {"观察仓", "加入自选"}
ADVICE_RED = {"放弃"}


def parse_pct(v):
    """'20.3%' / 0.203 -> 0.203 (float), 否则返回 None"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "").replace(",", "")
    try:
        return float(s) / 100.0
    except ValueError:
        return None


def parse_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def cell_value(raw, ctype):
    """返回 (显示文本, 数值或None, 是否百分比)"""
    if ctype == "pct":
        f = parse_pct(raw)
        if f is None:
            return ("", None, True)
        return (f"{f*100:.2f}%", f, True)
    if ctype in ("num2", "score"):
        f = parse_num(raw)
        if f is None:
            return ("", None, False)
        return (f"{f:.2f}", f, False)
    if ctype == "no":
        return (str(raw), None, False)
    return ("" if raw is None else str(raw), None, False)


def advice_color(advice):
    a = (advice or "").strip()
    if a in ADVICE_GREEN:
        return "FF0F6E56"
    if a in ADVICE_ORANGE:
        return "FFBA7517"
    if a in ADVICE_RED:
        return "FFA32D2D"
    return None


# ---------- HTML ----------
def build_html(meta, stocks):
    rows = []
    for s in stocks:
        tds = []
        for header, field, ctype in COLUMNS:
            val = s.get(field)
            text, _, _ = cell_value(val, ctype)
            style = ""
            if ctype == "score":
                style = ' style="text-align:center;"'
            if ctype == "advice":
                col = advice_color(val)
                if col:
                    style = f' style="color:{col};font-weight:600;text-align:center;"'
            tds.append(f"<td{style}>{text}</td>")
        rows.append("<tr>" + "".join(tds) + "</tr>")

    # 摘要区
    m = meta or {}
    ind_rows = "".join(
        f"<tr><td>{k}</td><td style='text-align:right;'>{v}</td></tr>"
        for k, v in (m.get("industry_distribution") or [])
    )
    top_rows = "".join(
        f"<tr><td>{i+1}</td><td>{t.get('code','')}</td><td>{t.get('name','')}</td>"
        f"<td style='text-align:right;'>{t.get('score', t.get('weighted_total',''))}</td></tr>"
        for i, t in enumerate(m.get("top10") or [])
    )
    exc = m.get("excluded_by") or {}
    exc_rows = "".join(
        f"<tr><td>否决项 {k}</td><td style='text-align:right;'>{v}</td></tr>"
        for k, v in exc.items()
    )
    ver = m.get("verification") or {}
    missing = m.get("missing_items") or "无"
    disclaimer = m.get("disclaimer") or "本结果仅基于量化模型，不构成投资建议。"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>巴菲特筛选_{m.get('trade_date','')}</title>
<style>
body{{font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;font-size:13px;color:#1a1a2e;margin:0;background:#f0f1f3;}}
.wrap{{max-width:1500px;margin:0 auto;padding:20px;}}
h1{{font-size:20px;margin:0 0 4px;}}
.sub{{color:#5a5a7a;margin-bottom:16px;font-size:12px;}}
.card{{background:#fff;border:1px solid rgba(0,0,0,.08);border-radius:12px;padding:16px 20px;margin-bottom:16px;}}
.card h2{{font-size:15px;margin:0 0 10px;}}
table{{border-collapse:collapse;width:100%;font-size:12px;}}
th,td{{border:1px solid rgba(0,0,0,.1);padding:6px 8px;text-align:left;}}
th{{background:#f0f1f3;font-weight:600;position:sticky;top:0;}}
.scroll{{overflow:auto;max-height:70vh;}}
.tag{{display:inline-block;padding:1px 6px;border-radius:4px;font-weight:600;}}
.green{{color:#0F6E56;}} .amber{{color:#BA7517;}} .red{{color:#A32D2D;}}
.kv{{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:8px;}}
.kv div{{background:#f8f9fa;border-radius:8px;padding:8px 10px;}}
.kv b{{display:block;font-size:18px;}}
.note{{font-size:12px;color:#5a5a7a;line-height:1.7;}}
</style></head><body><div class="wrap">
<h1>巴菲特安全边际法 · A股筛选报告</h1>
<div class="sub">筛选日期 {m.get('trade_date','')} ｜ 方法论：《巴菲特安全边际法 — 量化评估框架》v1.0</div>

<div class="card">
  <h2>摘要区</h2>
  <div class="kv">
    <div>全市场总数<b>{m.get('universe_count','-')}</b></div>
    <div>初筛通过数<b>{m.get('candidates_count','-')}</b></div>
    <div>最终分析数<b>{m.get('final_count', len(stocks))}</b></div>
    <div>加权总分最高<b>{stocks[0].get('weighted_total','-') if stocks else '-'}</b></div>
  </div>
</div>

<div class="card"><h2>各否决项排除数</h2><table>{exc_rows or '<tr><td>无</td></tr>'}</table></div>
<div class="card"><h2>行业板块分布</h2><table>{ind_rows or '<tr><td>无</td></tr>'}</table></div>
<div class="card"><h2>TOP 10 加权总分排行</h2><table><tr><th>排名</th><th>代码</th><th>名称</th><th>总分</th></tr>{top_rows or '<tr><td colspan=4>无</td></tr>'}</table></div>

<div class="card"><h2>筛选明细（{len(stocks)} 只）</h2>
<div class="scroll"><table>
<tr>{''.join(f'<th>{h}</th>' for h,_,_ in COLUMNS)}</tr>
{''.join(rows)}
</table></div></div>

<div class="card"><h2>三重复核执行状态</h2>
<div class="note">
<p><b>① 数据完整性：</b>{ver.get('data_completeness','未记录')}</p>
<p><b>② 评分一致性：</b>{ver.get('scoring_consistency','未记录')}</p>
<p><b>③ 决策合理性：</b>{ver.get('decision_rationality','未记录')}</p>
<p><b>数据缺失项：</b>{missing}</p>
</div></div>

<div class="card note">免责声明：{disclaimer}</div>
</div></body></html>"""
    return html


# ---------- XLSX ----------
def build_xlsx(meta, stocks, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
    HEADER_FONT = Font(bold=True, color="1A1A1A")
    THIN = Side(style="thin", color="BBBBBB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---- 明细表 ----
    ws = wb.active
    ws.title = "筛选明细"
    headers = [h for h, _, _ in COLUMNS]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"

    for s in stocks:
        row = []
        for _, field, ctype in COLUMNS:
            val = s.get(field)
            text, num, is_pct = cell_value(val, ctype)
            if ctype == "pct" and num is not None:
                row.append(num)
            elif ctype in ("num2", "score") and num is not None:
                row.append(num)
            else:
                row.append(text)
        ws.append(row)

    # 样式 + 格式
    for r in range(2, len(stocks) + 2):
        for c, (_, _, ctype) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if ctype == "pct":
                cell.number_format = "0.00%"
                cell.alignment = CENTER
            elif ctype in ("num2", "score"):
                cell.number_format = "0.00"
                if ctype == "score":
                    cell.alignment = CENTER
                else:
                    cell.alignment = Alignment(horizontal="right")
            elif ctype == "advice":
                col = advice_color(stocks[r - 2].get("advice"))
                cell.alignment = CENTER
                if col:
                    cell.font = Font(color=col, bold=True)
            elif ctype == "no":
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        # 数据完整度着色（最后一列）
        comp = ws.cell(row=r, column=len(headers))
        if "高度不确定" in str(comp.value):
            comp.font = Font(color="FFA32D2D")
        elif "部分缺失" in str(comp.value):
            comp.font = Font(color="FFBA7517")

    # 自适应列宽
    for c, (h, _, _) in enumerate(COLUMNS, start=1):
        maxlen = len(h)
        for r in range(2, len(stocks) + 2):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(maxlen + 3, 9), 40)

    # ---- 汇总表 ----
    m = meta or {}
    ws2 = wb.create_sheet("汇总")
    ws2.append(["巴菲特安全边际法 · A股筛选汇总"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append(["筛选日期", m.get("trade_date", "")])
    ws2.append(["全市场总数", m.get("universe_count", "-")])
    ws2.append(["初筛通过数", m.get("candidates_count", "-")])
    ws2.append(["最终分析数", m.get("final_count", len(stocks))])
    ws2.append([])
    ws2.append(["各否决项排除数"])
    ws2["A6"].font = Font(bold=True)
    for k, v in (m.get("excluded_by") or {}).items():
        ws2.append([f"否决项 {k}", v])
    ws2.append([])
    ws2.append(["行业板块分布"])
    r0 = ws2.max_row
    ws2.cell(row=r0, column=1).font = Font(bold=True)
    for k, v in (m.get("industry_distribution") or []):
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["TOP 10 加权总分排行"])
    r1 = ws2.max_row
    ws2.cell(row=r1, column=1).font = Font(bold=True)
    ws2.append(["排名", "代码", "名称", "总分"])
    for i, t in enumerate(m.get("top10") or [], start=1):
        ws2.append([i, t.get("code", ""), t.get("name", ""), t.get("score", t.get("weighted_total", ""))])
    ws2.append([])
    ver = m.get("verification") or {}
    ws2.append(["三重复核 - 数据完整性", ver.get("data_completeness", "未记录")])
    ws2.append(["三重复核 - 评分一致性", ver.get("scoring_consistency", "未记录")])
    ws2.append(["三重复核 - 决策合理性", ver.get("decision_rationality", "未记录")])
    ws2.append(["数据缺失项说明", m.get("missing_items", "无")])
    ws2.append(["免责声明", m.get("disclaimer", "本结果仅基于量化模型，不构成投资建议。")])
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output-dir", required=True)
    ap.add_argument("--date", required=True)
    args = ap.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)
    meta = data.get("meta", {})
    stocks = data.get("stocks", [])
    # 按加权总分降序并补序号
    def w(s):
        v = parse_num(s.get("weighted_total"))
        return v if v is not None else -1
    stocks.sort(key=w, reverse=True)
    for i, s in enumerate(stocks, start=1):
        s["no"] = i

    os.makedirs(args.output_dir, exist_ok=True)
    html_path = os.path.join(args.output_dir, f"巴菲特筛选_{args.date}.html")
    xlsx_path = os.path.join(args.output_dir, f"巴菲特筛选_{args.date}.xlsx")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(build_html(meta, stocks))
    build_xlsx(meta, stocks, xlsx_path)
    print(f"HTML -> {html_path}")
    print(f"XLSX -> {xlsx_path}")
    print(f"共 {len(stocks)} 只候选股")


if __name__ == "__main__":
    main()

"""
EXAMPLE _raw_2026-07-21.json:
{
  "meta": {
    "trade_date": "2026-07-21",
    "universe_count": 5200,
    "candidates_count": 318,
    "final_count": 318,
    "excluded_by": {"1": 120, "2": 45, "3": 60, "4": 200, "5": 10, "6": 8, "7": 15, "8": 30},
    "industry_distribution": [["白酒", 12], ["银行", 20]],
    "top10": [{"code":"600519.SH","name":"贵州茅台","score":9.2}],
    "verification": {
      "data_completeness": "2026-07-21 18:00 完成，每只候选股各维度≥3源",
      "scoring_consistency": "2026-07-21 18:30 抽验30%一致",
      "decision_rationality": "2026-07-21 19:00 高分股已填风险提示"
    },
    "missing_items": "部分股票质押比例数据缺失，已标注",
    "disclaimer": "本结果仅基于量化模型，不构成投资建议。"
  },
  "stocks": [
    {
      "code":"600519.SH","name":"贵州茅台","industry":"白酒",
      "roe_5y":"20.3%","gross_margin_5y":"91.2%",
      "pe_ttm":30.5,"peg":0.9,"debt_ratio":"21.0%","ocf_netprofit":1.15,
      "score_profit":10,"score_moat":10,"score_valuation":8,"score_solvency":10,"score_quality":10,
      "bonus_factor":1,"weighted_total":9.2,
      "intrinsic_value":1850.0,"price":1480.0,"margin_of_safety":"24.3%",
      "advice":"分批买入","risk_note":"估值仍不便宜；宏观需求走弱；产能释放不及预期",
      "data_completeness":"数据源充足"
    }
  ]
}
"""
